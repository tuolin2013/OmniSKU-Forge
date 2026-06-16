# E:\laotuo_project\OmniSKU-Forge\backend\app\api\core\services\knowledge_base.py
import os
import openpyxl

class ProductDatabase:
    def __init__(self):
        # 内存字典，用于极速检索，结构: {"猫静宝": {"产品名称": "猫静宝", "原料组成": "...", ...}}
        self.records = {}

    def init_from_directory(self, dir_path: str = "data"):
        """
        宽进严出：遍历指定目录下的所有 .xlsx 文件，强力吸纳所有带“产品名称”的合法 SKU。
        """
        if not os.path.exists(dir_path):
            print(f"⚠️ [知识库] 找不到目录: {dir_path}")
            return

        total_pet = 0
        total_tea = 0
        
        for filename in os.listdir(dir_path):
            if not filename.endswith(".xlsx") or filename.startswith("~$"):
                continue
                
            file_path = os.path.join(dir_path, filename)
            
            # 根据文件名打基底标签
            category = "tea" if "茶" in filename else "pet"

            try:
                # data_only=True 确保读取的是计算后的值，而不是 Excel 公式
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active

                # 获取第一行作为动态表头
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).strip() if cell.value else f"Unnamed_{cell.column}")

                # 寻找“产品名称”列所在的位置作为主键索引
                try:
                    name_index = headers.index("产品名称")
                except ValueError:
                    print(f"⚠️ [知识库] 文件 {filename} 未找到『产品名称』列，跳过解析。")
                    continue

                # 从第二行开始遍历数据
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # 以“产品名称”作为主键，如果为空则说明是空行，跳过
                    sku_name = str(row[name_index]).strip() if row[name_index] else ""
                    if not sku_name or sku_name == "None":
                        continue
                    
                    # 容错读取：打包成全量字典，即使其他字段全是空
                    row_data = {}
                    for idx, header in enumerate(headers):
                        if idx < len(row):
                            val = row[idx]
                            row_data[header] = str(val).strip() if val is not None else ""
                        else:
                            row_data[header] = ""
                    
                    # 强制注入系统级类目标签
                    row_data["__system_category__"] = category
                    
                    # 存入常驻内存字典
                    self.records[sku_name] = row_data
                    
                    # 统计计数
                    if category == "tea":
                        total_tea += 1
                    else:
                        total_pet += 1
                        
            except Exception as e:
                print(f"❌ [知识库] 读取文件 {filename} 失败: {e}")

        total = total_pet + total_tea
        print(f"✅ [知识库] 载入成功：总计 {total} 个，其中宠物({total_pet}), 茶饮({total_tea})")
        print(f"📋 当前内存常驻产品线: {list(self.records.keys())}")

    def init_from_excel(self, file_path: str):
        # 兼容老接口调用（如果外部传具体文件，直接降级处理）
        dir_name = os.path.dirname(file_path)
        self.init_from_directory(dir_name)

    def get_sku_info(self, sku_name: str) -> dict:
        """
        根据前端传来的 sku_name（如'猫静宝'）返回该 SKU 的全部列数据
        """
        return self.records.get(sku_name, {})

    def get_category_tree(self) -> list:
        """
        将字典数据聚合成前端 Cascader 需要的树状结构:
        [
            {
                "value": "pet",
                "label": "🐶 宠物营养保健",
                "children": [{"value": "产品名", "label": "产品名"}, ...]
            },
            {
                "value": "tea",
                "label": "🍵 养生茶饮",
                "children": [{"value": "产品名", "label": "产品名"}, ...]
            }
        ]
        """
        pet_children = []
        tea_children = []

        for sku_name, data in self.records.items():
            # 优先从 Excel 读取系统分类字段
            category = data.get("__system_category__", "").lower()
            
            # 容错逻辑：如果 Excel 没写分类，根据产品名关键词自动归类
            if not category:
                pet_keywords = ["宝", "消", "平", "没", "猫", "犬"]
                tea_keywords = ["茶", "莓", "饮"]
                if any(k in sku_name for k in pet_keywords):
                    category = "pet"
                elif any(k in sku_name for k in tea_keywords):
                    category = "tea"
                else:
                    # 默认归类为 pet (因为目前业务主要是宠物)
                    category = "pet"

            if category == "pet":
                pet_children.append({"value": sku_name, "label": sku_name})
            elif category == "tea":
                tea_children.append({"value": sku_name, "label": sku_name})

        result = []
        if pet_children:
            result.append({
                "value": "pet",
                "label": "🐶 宠物营养保健",
                "children": pet_children
            })
        if tea_children:
            result.append({
                "value": "tea",
                "label": "🍵 养生茶饮",
                "children": tea_children
            })

        return result

# 全局单例，在 main.py 的 lifespan 中被调用加载
product_db = ProductDatabase()
